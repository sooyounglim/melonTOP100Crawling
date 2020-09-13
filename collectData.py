import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
from urllib.request import urlretrieve
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

########################################################################################################################

menu = input("멜론차트 옵션을 입력하세요: ")

now = datetime.datetime.now()

try:
    wb = openpyxl.load_workbook("MelonChartSheet.xlsx")
    sheetList = wb.sheetnames
    for sheet in sheetList :
        if menu == sheet :
            wb.remove(wb[menu])
            break
    sheet = wb.create_sheet(menu)
    sheet.append(["현재시각", now])
    sheet.append(["순위", "앨범 사진 링크 주소", "곡 명", "아티스트 명", "앨범 명"])
    print("불러오기 완료")

except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = menu
    sheet.append(["현재시각", now])
    sheet.append(["순위", "앨범 사진 링크 주소", "곡 명", "아티스트 명", "앨범 명"])
    print("새로 파일을 만들었습니다")

########################################################################################################################

if menu == "realtime" :
    raw = requests.get("https://www.melon.com/chart/index.htm", headers={"User-Agent": "Mozilla/5.0"})
else :
    raw = requests.get("https://www.melon.com/chart/"+menu+"/index.htm", headers={"User-Agent": "Mozilla/5.0"})
html = BeautifulSoup(raw.text, 'html.parser')

########################################################################################################################

# 순위 : td > div.wrap.t_center > span.rank
# 앨범 사진 : div.wrap > a > img
# 곡명 : div.ellipsis.rank01 a
# 아티스트명 : div.ellipsis.rank02 > a:nth-of-type(1)
# 앨범명 : div.ellipsis.rank03 a
# 좋아요 수 : div.wrap span.cnt

########################################################################################################################

for cnt in range(1, 3) :
    if cnt == 1 :
        container = html.select('tr#lst50')
    else :
        container = html.select('tr#lst100')

    for c in container:
        rank = c.select_one('td > div.wrap.t_center > span.rank').text.strip()
        print(rank)

        song = c.select_one('div.ellipsis.rank01 a').text.strip()
        print(song)

        artistNum = 1
        artistString = ""
        while True:
            if c.select_one('div.ellipsis.rank02 > a:nth-of-type(' + str(artistNum) + ')') is not None:
                artistString += (c.select_one('div.ellipsis.rank02 > a:nth-of-type(' + str(artistNum) + ')').text.strip() + '/')
            else:
                artistString = artistString[:-1]
                print(artistString)
                break
            artistNum += 1

        album = c.select_one('div.ellipsis.rank03 a').text.strip()
        print(album)

        # like = c.select('div.wrap span.cnt')
        # print(like)

        img = c.select_one('div.wrap > a > img')
        imgSrc = img.attrs["src"]
        print(imgSrc)
        urlretrieve(imgSrc, 'image/'+menu+'/'+rank+'_'+song[:2]+'.png')

        sheet.append([int(rank), imgSrc, song, artistString, album])

        print("="*20)

wb.save("MelonChartSheet.xlsx")
